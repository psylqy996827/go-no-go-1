#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
This experiment was created using PsychoPy3 Experiment Builder (v2021.2.3),
    on 六月 09, 2024, at 21:43
If you publish work using this script the most relevant publication is:

    Peirce J, Gray JR, Simpson S, MacAskill M, Höchenberger R, Sogo H, Kastman E, Lindeløv JK. (2019) 
        PsychoPy2: Experiments in behavior made easy Behav Res 51: 195. 
        https://doi.org/10.3758/s13428-018-01193-y

"""

from __future__ import absolute_import, division

import psychopy
psychopy.useVersion('2021.2.3')


from psychopy import locale_setup
from psychopy import prefs
from psychopy import sound, gui, visual, core, data, event, logging, clock, colors
from psychopy.constants import (NOT_STARTED, STARTED, PLAYING, PAUSED,
                                STOPPED, FINISHED, PRESSED, RELEASED, FOREVER)

import numpy as np  # whole numpy lib is available, prepend 'np.'
from numpy import (sin, cos, tan, log, log10, pi, average,
                   sqrt, std, deg2rad, rad2deg, linspace, asarray)
from numpy.random import random, randint, normal, shuffle, choice as randchoice
import os  # handy system and path functions
import sys  # to get file system encoding

from psychopy.hardware import keyboard



# Ensure that relative paths start from the same directory as this script
_thisDir = os.path.dirname(os.path.abspath(__file__))
os.chdir(_thisDir)

# Store info about the experiment session
psychopyVersion = '2021.2.3'
expName = 'go no-go practice'  # from the Builder filename that created this script
expInfo = {'participant': '', 'session': '001'}
dlg = gui.DlgFromDict(dictionary=expInfo, sortKeys=False, title=expName)
if dlg.OK == False:
    core.quit()  # user pressed cancel
expInfo['date'] = data.getDateStr()  # add a simple timestamp
expInfo['expName'] = expName
expInfo['psychopyVersion'] = psychopyVersion

# Data file name stem = absolute path + name; later add .psyexp, .csv, .log, etc
filename = _thisDir + os.sep + u'data/%s_%s_%s' % (expInfo['participant'], expName, expInfo['date'])

# An ExperimentHandler isn't essential but helps with data saving
thisExp = data.ExperimentHandler(name=expName, version='',
    extraInfo=expInfo, runtimeInfo=None,
    originPath='C:\\Users\\16079\\Desktop\\go_nogo_practice\\go_nogo_practice\\go no-go practice_lastrun.py',
    savePickle=True, saveWideText=True,
    dataFileName=filename)
# save a log file for detail verbose info
logFile = logging.LogFile(filename+'.log', level=logging.EXP)
logging.console.setLevel(logging.WARNING)  # this outputs to the screen, not a file

endExpNow = False  # flag for 'escape' or other condition => quit the exp
frameTolerance = 0.001  # how close to onset before 'same' frame

# Start Code - component code to be run after the window creation

# Setup the Window
win = visual.Window(
    size=[1920, 1080], fullscr=True, screen=0, 
    winType='pyglet', allowGUI=False, allowStencil=False,
    monitor='testMonitor', color=[0,0,0], colorSpace='rgb',
    blendMode='avg', useFBO=True, 
    units='height')
# store frame rate of monitor if we can measure it
expInfo['frameRate'] = win.getActualFrameRate()
if expInfo['frameRate'] != None:
    frameDur = 1.0 / round(expInfo['frameRate'])
else:
    frameDur = 1.0 / 60.0  # could not measure, so guess

# Setup eyetracking
ioDevice = ioConfig = ioSession = ioServer = eyetracker = None

# create a default keyboard (e.g. to check for escape)
defaultKeyboard = keyboard.Keyboard()

# Initialize components for Routine "指导语1"
指导语1Clock = core.Clock()
text_1 = visual.TextStim(win=win, name='text_1',
    text='欢迎您参加本次实验。\n请先输入您配偶的名字以及ta的两个昵称\n明白后按任意键开始输入',
    font='Open Sans',
    pos=(0, 0), height=0.05, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=None, 
    languageStyle='LTR',
    depth=0.0);
key_resp_1 = keyboard.Keyboard()

# Initialize components for Routine "输入"
输入Clock = core.Clock()
textbox_name1 = visual.TextBox2(
     win, text='请输入名称', font='SimHei',
     pos=(0, 0.2),     letterHeight=0.05,
     size=(None, None), borderWidth=2.0,
     color='white', colorSpace='rgb',
     opacity=None,
     bold=False, italic=False,
     lineSpacing=1.0,
     padding=0.0,
     anchor='center',
     fillColor=None, borderColor=None,
     flipHoriz=False, flipVert=False,
     editable=True,
     name='textbox_name1',
     autoLog=True,
)
textbox_name2 = visual.TextBox2(
     win, text='请输入昵称1', font='SimHei',
     pos=(0, 0),     letterHeight=0.05,
     size=(None, None), borderWidth=2.0,
     color='white', colorSpace='rgb',
     opacity=None,
     bold=False, italic=False,
     lineSpacing=1.0,
     padding=0.0,
     anchor='center',
     fillColor=None, borderColor=None,
     flipHoriz=False, flipVert=False,
     editable=True,
     name='textbox_name2',
     autoLog=True,
)
textbox_name = visual.TextBox2(
     win, text='请输入昵称2', font='SimHei',
     pos=(0, -.2),     letterHeight=0.05,
     size=(None, None), borderWidth=2.0,
     color='white', colorSpace='rgb',
     opacity=None,
     bold=False, italic=False,
     lineSpacing=1.0,
     padding=0.0,
     anchor='center',
     fillColor=None, borderColor=None,
     flipHoriz=False, flipVert=False,
     editable=True,
     name='textbox_name',
     autoLog=True,
)
button = visual.ButtonStim(win, 
    text='OK', font='simhei',
    pos=(0, -.4),
    letterHeight=0.05,
    size=None, borderWidth=0.0,
    fillColor='darkgrey', borderColor=None,
    color='white', colorSpace='rgb',
    opacity=None,
    bold=True, italic=False,
    padding=None,
    anchor='center',
    name='button'
)
button.buttonClock = core.Clock()

# Initialize components for Routine "正式指导语"
正式指导语Clock = core.Clock()
text_8 = visual.TextStim(win=win, name='text_8',
    text='请注意！！！\n当屏幕中央出现积极词、您配偶的名字或昵称时，\n请快速且准确的按键盘空格键。\n而当出现消极词时，\n请不要按任何键。\n明白后按任意键进入练习。',
    font='Open Sans',
    pos=(0, 0), height=0.05, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=None, 
    languageStyle='LTR',
    depth=0.0);
key_resp_2 = keyboard.Keyboard()

# Initialize components for Routine "注视点"
注视点Clock = core.Clock()
polygon_1 = visual.ShapeStim(
    win=win, name='polygon_1', vertices='cross',
    size=(0.1, 0.1),
    ori=0.0, pos=(0, 0),
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=None, depth=0.0, interpolate=True)
rt1=0

# Initialize components for Routine "刺激1"
刺激1Clock = core.Clock()
stim1 = visual.TextStim(win=win, name='stim1',
    text='',
    font='Open Sans',
    pos=(0, 0), height=0.1, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=None, 
    languageStyle='LTR',
    depth=0.0);
key_resp_12 = keyboard.Keyboard()

# Initialize components for Routine "feedback1"
feedback1Clock = core.Clock()
feedback = ''
a = []
text_4 = visual.TextStim(win=win, name='text_4',
    text='',
    font='Open Sans',
    pos=(0, 0), height=0.2, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=None, 
    languageStyle='LTR',
    depth=-1.0);

# Initialize components for Routine "空屏1"
空屏1Clock = core.Clock()
polygon_2 = visual.ShapeStim(
    win=win, name='polygon_2',
    size=(0.5, 0.5), vertices='triangle',
    ori=0.0, pos=(0, 0),
    lineWidth=1.0,     colorSpace='rgb',  lineColor=(0.0000, 0.0000, 0.0000), fillColor=(0.0000, 0.0000, 0.0000),
    opacity=None, depth=0.0, interpolate=True)

# Initialize components for Routine "结束语1"
结束语1Clock = core.Clock()
text_2 = visual.TextStim(win=win, name='text_2',
    text='辛苦啦，请稍事休息~\n\n休息好后请按任意键进行下一个练习！',
    font='Open Sans',
    pos=(0, 0), height=0.05, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=None, 
    languageStyle='LTR',
    depth=0.0);
key_resp_6 = keyboard.Keyboard()

# Initialize components for Routine "指导语2"
指导语2Clock = core.Clock()
text_6 = visual.TextStim(win=win, name='text_6',
    text='请注意！！！\n这次是当出现消极词、您配偶的名字或昵称时，\n请快速且准确的按键盘空格键。\n而当出现积极词时，\n请不要按任何键。\n明白后按任意键进入实验。',
    font='Open Sans',
    pos=(0, 0), height=0.05, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=None, 
    languageStyle='LTR',
    depth=0.0);
key_resp = keyboard.Keyboard()

# Initialize components for Routine "注视点"
注视点Clock = core.Clock()
polygon_1 = visual.ShapeStim(
    win=win, name='polygon_1', vertices='cross',
    size=(0.1, 0.1),
    ori=0.0, pos=(0, 0),
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=None, depth=0.0, interpolate=True)
rt1=0

# Initialize components for Routine "刺激2"
刺激2Clock = core.Clock()
stim2 = visual.TextStim(win=win, name='stim2',
    text='',
    font='Open Sans',
    pos=(0, 0), height=0.1, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=None, 
    languageStyle='LTR',
    depth=0.0);
key_resp_13 = keyboard.Keyboard()

# Initialize components for Routine "feedback2"
feedback2Clock = core.Clock()
feedback = ''
a = []
text_7 = visual.TextStim(win=win, name='text_7',
    text='',
    font='Open Sans',
    pos=(0, 0), height=0.2, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=None, 
    languageStyle='LTR',
    depth=-1.0);

# Initialize components for Routine "空屏1"
空屏1Clock = core.Clock()
polygon_2 = visual.ShapeStim(
    win=win, name='polygon_2',
    size=(0.5, 0.5), vertices='triangle',
    ori=0.0, pos=(0, 0),
    lineWidth=1.0,     colorSpace='rgb',  lineColor=(0.0000, 0.0000, 0.0000), fillColor=(0.0000, 0.0000, 0.0000),
    opacity=None, depth=0.0, interpolate=True)

# Initialize components for Routine "结束语2"
结束语2Clock = core.Clock()
text_5 = visual.TextStim(win=win, name='text_5',
    text='感谢您的参与，实验到此结束！\n按任意键退出。',
    font='Open Sans',
    pos=(0, 0), height=0.05, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=None, 
    languageStyle='LTR',
    depth=0.0);
key_resp_4 = keyboard.Keyboard()

# Create some handy timers
globalClock = core.Clock()  # to track the time since experiment started
routineTimer = core.CountdownTimer()  # to track time remaining of each (non-slip) routine 

# ------Prepare to start Routine "指导语1"-------
continueRoutine = True
# update component parameters for each repeat
key_resp_1.keys = []
key_resp_1.rt = []
_key_resp_1_allKeys = []
# keep track of which components have finished
指导语1Components = [text_1, key_resp_1]
for thisComponent in 指导语1Components:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
指导语1Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "指导语1"-------
while continueRoutine:
    # get current time
    t = 指导语1Clock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=指导语1Clock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *text_1* updates
    if text_1.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        text_1.frameNStart = frameN  # exact frame index
        text_1.tStart = t  # local t and not account for scr refresh
        text_1.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(text_1, 'tStartRefresh')  # time at next scr refresh
        text_1.setAutoDraw(True)
    
    # *key_resp_1* updates
    waitOnFlip = False
    if key_resp_1.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        key_resp_1.frameNStart = frameN  # exact frame index
        key_resp_1.tStart = t  # local t and not account for scr refresh
        key_resp_1.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(key_resp_1, 'tStartRefresh')  # time at next scr refresh
        key_resp_1.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(key_resp_1.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(key_resp_1.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if key_resp_1.status == STARTED and not waitOnFlip:
        theseKeys = key_resp_1.getKeys(keyList=None, waitRelease=False)
        _key_resp_1_allKeys.extend(theseKeys)
        if len(_key_resp_1_allKeys):
            key_resp_1.keys = _key_resp_1_allKeys[-1].name  # just the last key pressed
            key_resp_1.rt = _key_resp_1_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in 指导语1Components:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "指导语1"-------
for thisComponent in 指导语1Components:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('text_1.started', text_1.tStartRefresh)
thisExp.addData('text_1.stopped', text_1.tStopRefresh)
# check responses
if key_resp_1.keys in ['', [], None]:  # No response was made
    key_resp_1.keys = None
thisExp.addData('key_resp_1.keys',key_resp_1.keys)
if key_resp_1.keys != None:  # we had a response
    thisExp.addData('key_resp_1.rt', key_resp_1.rt)
thisExp.addData('key_resp_1.started', key_resp_1.tStartRefresh)
thisExp.addData('key_resp_1.stopped', key_resp_1.tStopRefresh)
thisExp.nextEntry()
# the Routine "指导语1" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# ------Prepare to start Routine "输入"-------
continueRoutine = True
# update component parameters for each repeat
textbox_name1.reset()
textbox_name2.reset()
textbox_name.reset()
# -*- coding: utf-8 -*-
# 在每次实验开始时
from openpyxl import load_workbook

def button_pressed():
    # 加载现有的 Excel 文件
    wb = load_workbook('loop1_test.xlsx')
    ws = wb['Sheet1']  # 确保工作表名称正确

    # 读取 TextBox 中的文本
    if 'textbox_name1' in globals() and textbox_name1.text != '':
        ws['A12'] = textbox_name1.text  # 写入到 A12 单元格
    if 'textbox_name2' in globals() and textbox_name2.text != '':
        ws['A13'] = textbox_name2.text  # 写入到 A13 单元格
    if 'textbox_name' in globals() and textbox_name.text != '':
        ws['A14'] = textbox_name.text  # 写入到 A14 单元格

    # 保存更改
    wb.save('loop1_test.xlsx')
    wb.close()


    # 加载现有的 Excel 文件
    wb = load_workbook('loop2_test.xlsx')
    ws = wb['Sheet1']  # 确保工作表名称正确

    # 读取 TextBox 中的文本
    if 'textbox_name1' in globals() and textbox_name1.text != '':
        ws['A12'] = textbox_name1.text  # 写入到 A12 单元格
    if 'textbox_name2' in globals() and textbox_name2.text != '':
        ws['A13'] = textbox_name2.text  # 写入到 A13 单元格
    if 'textbox_name' in globals() and textbox_name.text != '':
        ws['A14'] = textbox_name.text  # 写入到 A14 单元格

    # 保存更改
    wb.save('loop2_test.xlsx')
    wb.close()




# keep track of which components have finished
输入Components = [textbox_name1, textbox_name2, textbox_name, button]
for thisComponent in 输入Components:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
输入Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "输入"-------
while continueRoutine:
    # get current time
    t = 输入Clock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=输入Clock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *textbox_name1* updates
    if textbox_name1.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        textbox_name1.frameNStart = frameN  # exact frame index
        textbox_name1.tStart = t  # local t and not account for scr refresh
        textbox_name1.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(textbox_name1, 'tStartRefresh')  # time at next scr refresh
        textbox_name1.setAutoDraw(True)
    
    # *textbox_name2* updates
    if textbox_name2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        textbox_name2.frameNStart = frameN  # exact frame index
        textbox_name2.tStart = t  # local t and not account for scr refresh
        textbox_name2.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(textbox_name2, 'tStartRefresh')  # time at next scr refresh
        textbox_name2.setAutoDraw(True)
    
    # *textbox_name* updates
    if textbox_name.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        textbox_name.frameNStart = frameN  # exact frame index
        textbox_name.tStart = t  # local t and not account for scr refresh
        textbox_name.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(textbox_name, 'tStartRefresh')  # time at next scr refresh
        textbox_name.setAutoDraw(True)
    
    # *button* updates
    if button.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
        # keep track of start time/frame for later
        button.frameNStart = frameN  # exact frame index
        button.tStart = t  # local t and not account for scr refresh
        button.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(button, 'tStartRefresh')  # time at next scr refresh
        button.setAutoDraw(True)
    if button.status == STARTED:
        # check whether button has been pressed
        if button.isClicked:
            if not button.wasClicked:
                button.timesOn.append(button.buttonClock.getTime()) # store time of first click
                button.timesOff.append(button.buttonClock.getTime()) # store time clicked until
            else:
                button.timesOff[-1] = button.buttonClock.getTime() # update time clicked until
            if not button.wasClicked:
                continueRoutine = False  # end routine when button is clicked
                button_pressed()
            button.wasClicked = True  # if button is still clicked next frame, it is not a new click
        else:
            button.wasClicked = False  # if button is clicked next frame, it is a new click
    else:
        button.wasClicked = False  # if button is clicked next frame, it is a new click
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in 输入Components:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "输入"-------
for thisComponent in 输入Components:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('textbox_name1.text',textbox_name1.text)
thisExp.addData('textbox_name1.started', textbox_name1.tStartRefresh)
thisExp.addData('textbox_name1.stopped', textbox_name1.tStopRefresh)
thisExp.addData('textbox_name2.text',textbox_name2.text)
thisExp.addData('textbox_name2.started', textbox_name2.tStartRefresh)
thisExp.addData('textbox_name2.stopped', textbox_name2.tStopRefresh)
thisExp.addData('textbox_name.text',textbox_name.text)
thisExp.addData('textbox_name.started', textbox_name.tStartRefresh)
thisExp.addData('textbox_name.stopped', textbox_name.tStopRefresh)
thisExp.addData('button.started', button.tStartRefresh)
thisExp.addData('button.stopped', button.tStopRefresh)
thisExp.addData('button.numClicks', button.numClicks)
if button.numClicks:
   thisExp.addData('button.timesOn', button.timesOn)
   thisExp.addData('button.timesOff', button.timesOff)
else:
   thisExp.addData('button.timesOn', "")
   thisExp.addData('button.timesOff', "")
# the Routine "输入" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# ------Prepare to start Routine "正式指导语"-------
continueRoutine = True
# update component parameters for each repeat
key_resp_2.keys = []
key_resp_2.rt = []
_key_resp_2_allKeys = []
# keep track of which components have finished
正式指导语Components = [text_8, key_resp_2]
for thisComponent in 正式指导语Components:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
正式指导语Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "正式指导语"-------
while continueRoutine:
    # get current time
    t = 正式指导语Clock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=正式指导语Clock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *text_8* updates
    if text_8.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        text_8.frameNStart = frameN  # exact frame index
        text_8.tStart = t  # local t and not account for scr refresh
        text_8.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(text_8, 'tStartRefresh')  # time at next scr refresh
        text_8.setAutoDraw(True)
    
    # *key_resp_2* updates
    waitOnFlip = False
    if key_resp_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        key_resp_2.frameNStart = frameN  # exact frame index
        key_resp_2.tStart = t  # local t and not account for scr refresh
        key_resp_2.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(key_resp_2, 'tStartRefresh')  # time at next scr refresh
        key_resp_2.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(key_resp_2.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(key_resp_2.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if key_resp_2.status == STARTED and not waitOnFlip:
        theseKeys = key_resp_2.getKeys(keyList=None, waitRelease=False)
        _key_resp_2_allKeys.extend(theseKeys)
        if len(_key_resp_2_allKeys):
            key_resp_2.keys = _key_resp_2_allKeys[-1].name  # just the last key pressed
            key_resp_2.rt = _key_resp_2_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in 正式指导语Components:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "正式指导语"-------
for thisComponent in 正式指导语Components:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('text_8.started', text_8.tStartRefresh)
thisExp.addData('text_8.stopped', text_8.tStopRefresh)
# check responses
if key_resp_2.keys in ['', [], None]:  # No response was made
    key_resp_2.keys = None
thisExp.addData('key_resp_2.keys',key_resp_2.keys)
if key_resp_2.keys != None:  # we had a response
    thisExp.addData('key_resp_2.rt', key_resp_2.rt)
thisExp.addData('key_resp_2.started', key_resp_2.tStartRefresh)
thisExp.addData('key_resp_2.stopped', key_resp_2.tStopRefresh)
thisExp.nextEntry()
# the Routine "正式指导语" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# set up handler to look after randomisation of conditions etc
block1 = data.TrialHandler(nReps=2.0, method='random', 
    extraInfo=expInfo, originPath=-1,
    trialList=data.importConditions('loop1_test.xlsx'),
    seed=None, name='block1')
thisExp.addLoop(block1)  # add the loop to the experiment
thisBlock1 = block1.trialList[0]  # so we can initialise stimuli with some values
# abbreviate parameter names if possible (e.g. rgb = thisBlock1.rgb)
if thisBlock1 != None:
    for paramName in thisBlock1:
        exec('{} = thisBlock1[paramName]'.format(paramName))

for thisBlock1 in block1:
    currentLoop = block1
    # abbreviate parameter names if possible (e.g. rgb = thisBlock1.rgb)
    if thisBlock1 != None:
        for paramName in thisBlock1:
            exec('{} = thisBlock1[paramName]'.format(paramName))
    
    # ------Prepare to start Routine "注视点"-------
    continueRoutine = True
    routineTimer.add(0.500000)
    # update component parameters for each repeat
    from numpy.random import shuffle
    
    d=[0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
    shuffle(d)
    rt1=d.pop()
    
    thisExp.addData("rt1",rt1)
    
    # keep track of which components have finished
    注视点Components = [polygon_1]
    for thisComponent in 注视点Components:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    注视点Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "注视点"-------
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = 注视点Clock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=注视点Clock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *polygon_1* updates
        if polygon_1.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            polygon_1.frameNStart = frameN  # exact frame index
            polygon_1.tStart = t  # local t and not account for scr refresh
            polygon_1.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(polygon_1, 'tStartRefresh')  # time at next scr refresh
            polygon_1.setAutoDraw(True)
        if polygon_1.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > polygon_1.tStartRefresh + 0.5-frameTolerance:
                # keep track of stop time/frame for later
                polygon_1.tStop = t  # not accounting for scr refresh
                polygon_1.frameNStop = frameN  # exact frame index
                win.timeOnFlip(polygon_1, 'tStopRefresh')  # time at next scr refresh
                polygon_1.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in 注视点Components:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "注视点"-------
    for thisComponent in 注视点Components:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    block1.addData('polygon_1.started', polygon_1.tStartRefresh)
    block1.addData('polygon_1.stopped', polygon_1.tStopRefresh)
    
    # ------Prepare to start Routine "刺激1"-------
    continueRoutine = True
    routineTimer.add(1.000000)
    # update component parameters for each repeat
    stim1.setText(stimuli1)
    key_resp_12.keys = []
    key_resp_12.rt = []
    _key_resp_12_allKeys = []
    # keep track of which components have finished
    刺激1Components = [stim1, key_resp_12]
    for thisComponent in 刺激1Components:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    刺激1Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "刺激1"-------
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = 刺激1Clock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=刺激1Clock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *stim1* updates
        if stim1.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            stim1.frameNStart = frameN  # exact frame index
            stim1.tStart = t  # local t and not account for scr refresh
            stim1.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(stim1, 'tStartRefresh')  # time at next scr refresh
            stim1.setAutoDraw(True)
        if stim1.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > stim1.tStartRefresh + 0.6-frameTolerance:
                # keep track of stop time/frame for later
                stim1.tStop = t  # not accounting for scr refresh
                stim1.frameNStop = frameN  # exact frame index
                win.timeOnFlip(stim1, 'tStopRefresh')  # time at next scr refresh
                stim1.setAutoDraw(False)
        
        # *key_resp_12* updates
        waitOnFlip = False
        if key_resp_12.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            key_resp_12.frameNStart = frameN  # exact frame index
            key_resp_12.tStart = t  # local t and not account for scr refresh
            key_resp_12.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(key_resp_12, 'tStartRefresh')  # time at next scr refresh
            key_resp_12.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(key_resp_12.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(key_resp_12.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if key_resp_12.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > key_resp_12.tStartRefresh + 1.0-frameTolerance:
                # keep track of stop time/frame for later
                key_resp_12.tStop = t  # not accounting for scr refresh
                key_resp_12.frameNStop = frameN  # exact frame index
                win.timeOnFlip(key_resp_12, 'tStopRefresh')  # time at next scr refresh
                key_resp_12.status = FINISHED
        if key_resp_12.status == STARTED and not waitOnFlip:
            theseKeys = key_resp_12.getKeys(keyList=['space'], waitRelease=False)
            _key_resp_12_allKeys.extend(theseKeys)
            if len(_key_resp_12_allKeys):
                key_resp_12.keys = _key_resp_12_allKeys[-1].name  # just the last key pressed
                key_resp_12.rt = _key_resp_12_allKeys[-1].rt
                # was this correct?
                if (key_resp_12.keys == str(correct)) or (key_resp_12.keys == correct):
                    key_resp_12.corr = 1
                else:
                    key_resp_12.corr = 0
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in 刺激1Components:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "刺激1"-------
    for thisComponent in 刺激1Components:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    block1.addData('stim1.started', stim1.tStartRefresh)
    block1.addData('stim1.stopped', stim1.tStopRefresh)
    # check responses
    if key_resp_12.keys in ['', [], None]:  # No response was made
        key_resp_12.keys = None
        # was no response the correct answer?!
        if str(correct).lower() == 'none':
           key_resp_12.corr = 1;  # correct non-response
        else:
           key_resp_12.corr = 0;  # failed to respond (incorrectly)
    # store data for block1 (TrialHandler)
    block1.addData('key_resp_12.keys',key_resp_12.keys)
    block1.addData('key_resp_12.corr', key_resp_12.corr)
    if key_resp_12.keys != None:  # we had a response
        block1.addData('key_resp_12.rt', key_resp_12.rt)
    block1.addData('key_resp_12.started', key_resp_12.tStartRefresh)
    block1.addData('key_resp_12.stopped', key_resp_12.tStopRefresh)
    
    # ------Prepare to start Routine "feedback1"-------
    continueRoutine = True
    routineTimer.add(0.100000)
    # update component parameters for each repeat
    
    # 假设 key_resp_12.keys 是按键响应, key_resp_7.corr 是一个布尔值，表示响应是否正确
    if signal == 'Go':
        if key_resp_12.keys == 'space':
            feedback = "正确!"
            feedback_color = 'green'
            key_resp_12.corr = True
        else:
            feedback = "错误!" 
            feedback_color = 'red'
            key_resp_12.corr = False
        
    elif signal == 'No-Go':
        if key_resp_12.keys is None or key_resp_12.keys == '':
            feedback = "正确!"
            feedback_color = 'green'
            key_resp_12.corr = True
        else:
            feedback = "错误!" 
            feedback_color = 'red'
            key_resp_12.corr = False
        
    # 累计正确响应次数
    if 'a' not in locals():  # 初始化a
        a = []
    
    a.append(key_resp_12.corr)
    b = sum(a)
    
    
    
    
    
     
    text_4.setColor(feedback_color, colorSpace='rgb')
    text_4.setText(feedback)
    # keep track of which components have finished
    feedback1Components = [text_4]
    for thisComponent in feedback1Components:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    feedback1Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "feedback1"-------
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = feedback1Clock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=feedback1Clock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *text_4* updates
        if text_4.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            text_4.frameNStart = frameN  # exact frame index
            text_4.tStart = t  # local t and not account for scr refresh
            text_4.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(text_4, 'tStartRefresh')  # time at next scr refresh
            text_4.setAutoDraw(True)
        if text_4.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > text_4.tStartRefresh + 0.1-frameTolerance:
                # keep track of stop time/frame for later
                text_4.tStop = t  # not accounting for scr refresh
                text_4.frameNStop = frameN  # exact frame index
                win.timeOnFlip(text_4, 'tStopRefresh')  # time at next scr refresh
                text_4.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in feedback1Components:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "feedback1"-------
    for thisComponent in feedback1Components:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    block1.addData('text_4.started', text_4.tStartRefresh)
    block1.addData('text_4.stopped', text_4.tStopRefresh)
    
    # ------Prepare to start Routine "空屏1"-------
    continueRoutine = True
    routineTimer.add(0.400000)
    # update component parameters for each repeat
    # keep track of which components have finished
    空屏1Components = [polygon_2]
    for thisComponent in 空屏1Components:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    空屏1Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "空屏1"-------
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = 空屏1Clock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=空屏1Clock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *polygon_2* updates
        if polygon_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            polygon_2.frameNStart = frameN  # exact frame index
            polygon_2.tStart = t  # local t and not account for scr refresh
            polygon_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(polygon_2, 'tStartRefresh')  # time at next scr refresh
            polygon_2.setAutoDraw(True)
        if polygon_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > polygon_2.tStartRefresh + 0.4-frameTolerance:
                # keep track of stop time/frame for later
                polygon_2.tStop = t  # not accounting for scr refresh
                polygon_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(polygon_2, 'tStopRefresh')  # time at next scr refresh
                polygon_2.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in 空屏1Components:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "空屏1"-------
    for thisComponent in 空屏1Components:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    block1.addData('polygon_2.started', polygon_2.tStartRefresh)
    block1.addData('polygon_2.stopped', polygon_2.tStopRefresh)
    thisExp.nextEntry()
    
# completed 2.0 repeats of 'block1'


# ------Prepare to start Routine "结束语1"-------
continueRoutine = True
# update component parameters for each repeat
key_resp_6.keys = []
key_resp_6.rt = []
_key_resp_6_allKeys = []
# keep track of which components have finished
结束语1Components = [text_2, key_resp_6]
for thisComponent in 结束语1Components:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
结束语1Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "结束语1"-------
while continueRoutine:
    # get current time
    t = 结束语1Clock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=结束语1Clock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *text_2* updates
    if text_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        text_2.frameNStart = frameN  # exact frame index
        text_2.tStart = t  # local t and not account for scr refresh
        text_2.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(text_2, 'tStartRefresh')  # time at next scr refresh
        text_2.setAutoDraw(True)
    
    # *key_resp_6* updates
    waitOnFlip = False
    if key_resp_6.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        key_resp_6.frameNStart = frameN  # exact frame index
        key_resp_6.tStart = t  # local t and not account for scr refresh
        key_resp_6.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(key_resp_6, 'tStartRefresh')  # time at next scr refresh
        key_resp_6.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(key_resp_6.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(key_resp_6.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if key_resp_6.status == STARTED and not waitOnFlip:
        theseKeys = key_resp_6.getKeys(keyList=None, waitRelease=False)
        _key_resp_6_allKeys.extend(theseKeys)
        if len(_key_resp_6_allKeys):
            key_resp_6.keys = _key_resp_6_allKeys[-1].name  # just the last key pressed
            key_resp_6.rt = _key_resp_6_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in 结束语1Components:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "结束语1"-------
for thisComponent in 结束语1Components:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('text_2.started', text_2.tStartRefresh)
thisExp.addData('text_2.stopped', text_2.tStopRefresh)
# check responses
if key_resp_6.keys in ['', [], None]:  # No response was made
    key_resp_6.keys = None
thisExp.addData('key_resp_6.keys',key_resp_6.keys)
if key_resp_6.keys != None:  # we had a response
    thisExp.addData('key_resp_6.rt', key_resp_6.rt)
thisExp.addData('key_resp_6.started', key_resp_6.tStartRefresh)
thisExp.addData('key_resp_6.stopped', key_resp_6.tStopRefresh)
thisExp.nextEntry()
# the Routine "结束语1" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# ------Prepare to start Routine "指导语2"-------
continueRoutine = True
# update component parameters for each repeat
key_resp.keys = []
key_resp.rt = []
_key_resp_allKeys = []
# keep track of which components have finished
指导语2Components = [text_6, key_resp]
for thisComponent in 指导语2Components:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
指导语2Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "指导语2"-------
while continueRoutine:
    # get current time
    t = 指导语2Clock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=指导语2Clock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *text_6* updates
    if text_6.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        text_6.frameNStart = frameN  # exact frame index
        text_6.tStart = t  # local t and not account for scr refresh
        text_6.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(text_6, 'tStartRefresh')  # time at next scr refresh
        text_6.setAutoDraw(True)
    
    # *key_resp* updates
    waitOnFlip = False
    if key_resp.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        key_resp.frameNStart = frameN  # exact frame index
        key_resp.tStart = t  # local t and not account for scr refresh
        key_resp.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(key_resp, 'tStartRefresh')  # time at next scr refresh
        key_resp.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(key_resp.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(key_resp.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if key_resp.status == STARTED and not waitOnFlip:
        theseKeys = key_resp.getKeys(keyList=None, waitRelease=False)
        _key_resp_allKeys.extend(theseKeys)
        if len(_key_resp_allKeys):
            key_resp.keys = _key_resp_allKeys[-1].name  # just the last key pressed
            key_resp.rt = _key_resp_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in 指导语2Components:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "指导语2"-------
for thisComponent in 指导语2Components:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('text_6.started', text_6.tStartRefresh)
thisExp.addData('text_6.stopped', text_6.tStopRefresh)
# check responses
if key_resp.keys in ['', [], None]:  # No response was made
    key_resp.keys = None
thisExp.addData('key_resp.keys',key_resp.keys)
if key_resp.keys != None:  # we had a response
    thisExp.addData('key_resp.rt', key_resp.rt)
thisExp.addData('key_resp.started', key_resp.tStartRefresh)
thisExp.addData('key_resp.stopped', key_resp.tStopRefresh)
thisExp.nextEntry()
# the Routine "指导语2" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# set up handler to look after randomisation of conditions etc
block2 = data.TrialHandler(nReps=2.0, method='random', 
    extraInfo=expInfo, originPath=-1,
    trialList=data.importConditions('loop2_test.xlsx'),
    seed=None, name='block2')
thisExp.addLoop(block2)  # add the loop to the experiment
thisBlock2 = block2.trialList[0]  # so we can initialise stimuli with some values
# abbreviate parameter names if possible (e.g. rgb = thisBlock2.rgb)
if thisBlock2 != None:
    for paramName in thisBlock2:
        exec('{} = thisBlock2[paramName]'.format(paramName))

for thisBlock2 in block2:
    currentLoop = block2
    # abbreviate parameter names if possible (e.g. rgb = thisBlock2.rgb)
    if thisBlock2 != None:
        for paramName in thisBlock2:
            exec('{} = thisBlock2[paramName]'.format(paramName))
    
    # ------Prepare to start Routine "注视点"-------
    continueRoutine = True
    routineTimer.add(0.500000)
    # update component parameters for each repeat
    from numpy.random import shuffle
    
    d=[0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
    shuffle(d)
    rt1=d.pop()
    
    thisExp.addData("rt1",rt1)
    
    # keep track of which components have finished
    注视点Components = [polygon_1]
    for thisComponent in 注视点Components:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    注视点Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "注视点"-------
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = 注视点Clock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=注视点Clock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *polygon_1* updates
        if polygon_1.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            polygon_1.frameNStart = frameN  # exact frame index
            polygon_1.tStart = t  # local t and not account for scr refresh
            polygon_1.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(polygon_1, 'tStartRefresh')  # time at next scr refresh
            polygon_1.setAutoDraw(True)
        if polygon_1.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > polygon_1.tStartRefresh + 0.5-frameTolerance:
                # keep track of stop time/frame for later
                polygon_1.tStop = t  # not accounting for scr refresh
                polygon_1.frameNStop = frameN  # exact frame index
                win.timeOnFlip(polygon_1, 'tStopRefresh')  # time at next scr refresh
                polygon_1.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in 注视点Components:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "注视点"-------
    for thisComponent in 注视点Components:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    block2.addData('polygon_1.started', polygon_1.tStartRefresh)
    block2.addData('polygon_1.stopped', polygon_1.tStopRefresh)
    
    # ------Prepare to start Routine "刺激2"-------
    continueRoutine = True
    routineTimer.add(1.000000)
    # update component parameters for each repeat
    stim2.setText(stimuli2)
    key_resp_13.keys = []
    key_resp_13.rt = []
    _key_resp_13_allKeys = []
    # keep track of which components have finished
    刺激2Components = [stim2, key_resp_13]
    for thisComponent in 刺激2Components:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    刺激2Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "刺激2"-------
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = 刺激2Clock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=刺激2Clock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *stim2* updates
        if stim2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            stim2.frameNStart = frameN  # exact frame index
            stim2.tStart = t  # local t and not account for scr refresh
            stim2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(stim2, 'tStartRefresh')  # time at next scr refresh
            stim2.setAutoDraw(True)
        if stim2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > stim2.tStartRefresh + 0.6-frameTolerance:
                # keep track of stop time/frame for later
                stim2.tStop = t  # not accounting for scr refresh
                stim2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(stim2, 'tStopRefresh')  # time at next scr refresh
                stim2.setAutoDraw(False)
        
        # *key_resp_13* updates
        waitOnFlip = False
        if key_resp_13.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            key_resp_13.frameNStart = frameN  # exact frame index
            key_resp_13.tStart = t  # local t and not account for scr refresh
            key_resp_13.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(key_resp_13, 'tStartRefresh')  # time at next scr refresh
            key_resp_13.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(key_resp_13.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(key_resp_13.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if key_resp_13.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > key_resp_13.tStartRefresh + 1.0-frameTolerance:
                # keep track of stop time/frame for later
                key_resp_13.tStop = t  # not accounting for scr refresh
                key_resp_13.frameNStop = frameN  # exact frame index
                win.timeOnFlip(key_resp_13, 'tStopRefresh')  # time at next scr refresh
                key_resp_13.status = FINISHED
        if key_resp_13.status == STARTED and not waitOnFlip:
            theseKeys = key_resp_13.getKeys(keyList=['space'], waitRelease=False)
            _key_resp_13_allKeys.extend(theseKeys)
            if len(_key_resp_13_allKeys):
                key_resp_13.keys = _key_resp_13_allKeys[-1].name  # just the last key pressed
                key_resp_13.rt = _key_resp_13_allKeys[-1].rt
                # was this correct?
                if (key_resp_13.keys == str(correct)) or (key_resp_13.keys == correct):
                    key_resp_13.corr = 1
                else:
                    key_resp_13.corr = 0
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in 刺激2Components:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "刺激2"-------
    for thisComponent in 刺激2Components:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    block2.addData('stim2.started', stim2.tStartRefresh)
    block2.addData('stim2.stopped', stim2.tStopRefresh)
    # check responses
    if key_resp_13.keys in ['', [], None]:  # No response was made
        key_resp_13.keys = None
        # was no response the correct answer?!
        if str(correct).lower() == 'none':
           key_resp_13.corr = 1;  # correct non-response
        else:
           key_resp_13.corr = 0;  # failed to respond (incorrectly)
    # store data for block2 (TrialHandler)
    block2.addData('key_resp_13.keys',key_resp_13.keys)
    block2.addData('key_resp_13.corr', key_resp_13.corr)
    if key_resp_13.keys != None:  # we had a response
        block2.addData('key_resp_13.rt', key_resp_13.rt)
    block2.addData('key_resp_13.started', key_resp_13.tStartRefresh)
    block2.addData('key_resp_13.stopped', key_resp_13.tStopRefresh)
    
    # ------Prepare to start Routine "feedback2"-------
    continueRoutine = True
    routineTimer.add(0.100000)
    # update component parameters for each repeat
    # 假设 key_resp_13.keys 是按键响应, key_resp_7.corr 是一个布尔值，表示响应是否正确
    if signal == 'Go':
        if key_resp_13.keys == 'space':
            feedback = "正确!"
            feedback_color = 'green'
            key_resp_13.corr = True
        else:
            feedback = "错误!" 
            feedback_color = 'red'
            key_resp_13.corr = False
        
    elif signal == 'No-Go':
        if key_resp_13.keys is None or key_resp_13.keys == '':
            feedback = "正确!"
            feedback_color = 'green'
            key_resp_13.corr = True
        else:
            feedback = "错误!" 
            feedback_color = 'red'
            key_resp_13.corr = False
        
    # 累计正确响应次数
    if 'a' not in locals():  # 初始化a
        a = []
    
    a.append(key_resp_13.corr)
    b = sum(a)
    
    
    
    
     
    text_7.setColor(feedback_color, colorSpace='rgb')
    text_7.setText(feedback)
    # keep track of which components have finished
    feedback2Components = [text_7]
    for thisComponent in feedback2Components:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    feedback2Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "feedback2"-------
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = feedback2Clock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=feedback2Clock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *text_7* updates
        if text_7.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            text_7.frameNStart = frameN  # exact frame index
            text_7.tStart = t  # local t and not account for scr refresh
            text_7.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(text_7, 'tStartRefresh')  # time at next scr refresh
            text_7.setAutoDraw(True)
        if text_7.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > text_7.tStartRefresh + 0.1-frameTolerance:
                # keep track of stop time/frame for later
                text_7.tStop = t  # not accounting for scr refresh
                text_7.frameNStop = frameN  # exact frame index
                win.timeOnFlip(text_7, 'tStopRefresh')  # time at next scr refresh
                text_7.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in feedback2Components:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "feedback2"-------
    for thisComponent in feedback2Components:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    block2.addData('text_7.started', text_7.tStartRefresh)
    block2.addData('text_7.stopped', text_7.tStopRefresh)
    
    # ------Prepare to start Routine "空屏1"-------
    continueRoutine = True
    routineTimer.add(0.400000)
    # update component parameters for each repeat
    # keep track of which components have finished
    空屏1Components = [polygon_2]
    for thisComponent in 空屏1Components:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    空屏1Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "空屏1"-------
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = 空屏1Clock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=空屏1Clock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *polygon_2* updates
        if polygon_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            polygon_2.frameNStart = frameN  # exact frame index
            polygon_2.tStart = t  # local t and not account for scr refresh
            polygon_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(polygon_2, 'tStartRefresh')  # time at next scr refresh
            polygon_2.setAutoDraw(True)
        if polygon_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > polygon_2.tStartRefresh + 0.4-frameTolerance:
                # keep track of stop time/frame for later
                polygon_2.tStop = t  # not accounting for scr refresh
                polygon_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(polygon_2, 'tStopRefresh')  # time at next scr refresh
                polygon_2.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in 空屏1Components:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "空屏1"-------
    for thisComponent in 空屏1Components:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    block2.addData('polygon_2.started', polygon_2.tStartRefresh)
    block2.addData('polygon_2.stopped', polygon_2.tStopRefresh)
    thisExp.nextEntry()
    
# completed 2.0 repeats of 'block2'


# ------Prepare to start Routine "结束语2"-------
continueRoutine = True
# update component parameters for each repeat
key_resp_4.keys = []
key_resp_4.rt = []
_key_resp_4_allKeys = []
# keep track of which components have finished
结束语2Components = [text_5, key_resp_4]
for thisComponent in 结束语2Components:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
结束语2Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "结束语2"-------
while continueRoutine:
    # get current time
    t = 结束语2Clock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=结束语2Clock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *text_5* updates
    if text_5.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        text_5.frameNStart = frameN  # exact frame index
        text_5.tStart = t  # local t and not account for scr refresh
        text_5.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(text_5, 'tStartRefresh')  # time at next scr refresh
        text_5.setAutoDraw(True)
    
    # *key_resp_4* updates
    waitOnFlip = False
    if key_resp_4.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        key_resp_4.frameNStart = frameN  # exact frame index
        key_resp_4.tStart = t  # local t and not account for scr refresh
        key_resp_4.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(key_resp_4, 'tStartRefresh')  # time at next scr refresh
        key_resp_4.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(key_resp_4.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(key_resp_4.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if key_resp_4.status == STARTED and not waitOnFlip:
        theseKeys = key_resp_4.getKeys(keyList=None, waitRelease=False)
        _key_resp_4_allKeys.extend(theseKeys)
        if len(_key_resp_4_allKeys):
            key_resp_4.keys = _key_resp_4_allKeys[-1].name  # just the last key pressed
            key_resp_4.rt = _key_resp_4_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in 结束语2Components:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "结束语2"-------
for thisComponent in 结束语2Components:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('text_5.started', text_5.tStartRefresh)
thisExp.addData('text_5.stopped', text_5.tStopRefresh)
# check responses
if key_resp_4.keys in ['', [], None]:  # No response was made
    key_resp_4.keys = None
thisExp.addData('key_resp_4.keys',key_resp_4.keys)
if key_resp_4.keys != None:  # we had a response
    thisExp.addData('key_resp_4.rt', key_resp_4.rt)
thisExp.addData('key_resp_4.started', key_resp_4.tStartRefresh)
thisExp.addData('key_resp_4.stopped', key_resp_4.tStopRefresh)
thisExp.nextEntry()
# the Routine "结束语2" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# Flip one final time so any remaining win.callOnFlip() 
# and win.timeOnFlip() tasks get executed before quitting
win.flip()

# these shouldn't be strictly necessary (should auto-save)
thisExp.saveAsWideText(filename+'.csv', delim='auto')
thisExp.saveAsPickle(filename)
logging.flush()
# make sure everything is closed down
thisExp.abort()  # or data files will save again on exit
win.close()
core.quit()
